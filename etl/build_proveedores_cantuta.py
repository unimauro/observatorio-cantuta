import openpyxl, json, os
HERE=os.path.dirname(os.path.abspath(__file__))
CONOSCE=os.environ.get('CONOSCE_DIR','/Users/unimauro/Documents/Repos/observatorio-sanmarcos/etl/conosce')
CODIGO='001901'; RUC_ENT='20174950971'  # UNE La Cantuta - pliego 528
def tp(ruc):
    s=str(ruc)
    # consorcio/codigo SEACE: no 11 digitos o no empieza en 10/15/20 -> empresa
    if len(s)!=11 or s[:2] not in ('10','15','20'):
        return 'empresa'
    return 'natural' if s[:2] in ('10','15') else 'empresa'

# dueno/representante (via datosperu.org / reporte consorcios CONOSCE). null = sin fuente confiable.
DP='https://www.datosperu.org/'
CONS='https://conosce.osce.gob.pe/buscador/assets/67ae6c4a/reportes/consorcios/'
duenos={
 # --- CONSORCIOS (codigo SEACE) -> integrantes y sus gerentes (Reporte de Consorcios CONOSCE 2025) ---
 '1647807':('Consorcio integrado por GRANDES XTRUCTURAS S.A.C. (RUC 20605238468, gte. Zea Rojas Christoph Willy), GONSAGI INGENIEROS S.A.C. (RUC 20600482506, gte. Gonzales Sanchez Gilmer) y MAQUINARIA Y CONSTRUCCION CONDOR ANDINO E.I.R.L. (RUC 20607483907, titular-gte. Benavides Cieza Geiner)', CONS+'2025/CONOSCE_CONSORCIO2025_0.xlsx'),
 '1740854':('Consorcio integrado por GRUPO G & S SECURITY S.A.C. (RUC 20604499748, gte. Medina Huerto Yemmy Gilmer) y SEGURIDAD JUCINU SELVA S.A.C. (RUC 20606346639, gte. Franco Fuentes Micaela Julia)', CONS+'2025/CONOSCE_CONSORCIO2025_0.xlsx'),
 '1693250':('Consorcio integrado por GRUPO G & S SECURITY S.A.C. (RUC 20604499748, gte. Medina Huerto Yemmy Gilmer) y SEGURIDAD JUCINU SELVA S.A.C. (RUC 20606346639, gte. Franco Fuentes Micaela Julia)', CONS+'2025/CONOSCE_CONSORCIO2025_0.xlsx'),
 # --- EMPRESAS (RUC 20) -> gerente general / representante segun datosperu.org ---
 '20416207900':('Gerente general: Chong Muñoz Juan Jose (director: Schmidt Alvarez Emilio Alberto)', DP+'empresa-transportes-thornado-sac-20416207900.php'),
 '20330055805':('Titular-gerente: Vasquez Zavala Jose Teofilo', DP+'empresa-distribuidora-jovaza-eirl-20330055805.php'),
 '20608794388':('Gerente general: Alcarraz Mucha Aderly Franks', DP+'empresa-goods-island-partners-sac-20608794388.php'),
 '20568286562':('Gerente general: Minaya Daza Hillary Jenyfer (C.I.S. Viproser SAC)', DP+'empresa-compania-de-inversiones-y-servicios-vigilancia-proteccion-seguridad-y-resguardo-sac-20568286562.php'),
 '20100070970':('Supermercados Peruanos S.A. (Plaza Vea/Vivanda); grupo Intercorp Retail / InRetail, controlado por Carlos Rodriguez-Pastor Persivale', DP+'empresa-supermercados-peruanos-sociedad-anonima-o-s-p-s-a-20100070970.php'),
 '20600270002':('Gerente: Macedo Diaz Frank (apoderado: Renteria Osorio Rodrigo Fabian)', DP+'empresa-rgr-equipos-y-maquinarias-eirl-20600270002.php'),
 '20109072177':('Cencosud Retail Peru S.A. (Wong/Metro); grupo Cencosud S.A. (Chile, familia Paulmann)', DP+'empresa-cencosud-retail-peru-sa-20109072177.php'),
 '20101320295':('Director gerente: Cardenas Acosta Amado Manuel (Laboratorio Diesel Senatinos S.A.)', DP+'empresa-laboratorio-diesel-senatinos-sa-20101320295.php'),
 '20523550871':('Gerente general: Abregu Castro Denissi Yanina', DP+'empresa-negociaciones-caryder-sac-20523550871.php'),
}

rows=[]
for Y in (2023,2024,2025):
    wb=openpyxl.load_workbook(f"{CONOSCE}/CONOSCE_ADJUDICACIONES{Y}_0.xlsx", read_only=True)
    ws=wb[wb.sheetnames[0]]
    it=ws.iter_rows(values_only=True); next(it)
    for r in it:
        if str(r[0])==CODIGO and str(r[1])==RUC_ENT:
            rows.append(r)
    wb.close()
print("item rows:", len(rows))

agg={}
for r in rows:
    ruc=str(r[19]); monto=r[15] or 0
    a=agg.setdefault(ruc,{'nombre':r[20],'ruc':ruc,'monto':0.0,'convs':set(),
        'tipos':{},'objeto':{},'tipo_persona':tp(ruc),'tipo_proveedor_seace':r[21]})
    a['monto']+=float(monto)
    a['convs'].add(r[5])
    a['tipos'][r[7]]=a['tipos'].get(r[7],0)+1
    a['objeto'][r[6]]=a['objeto'].get(r[6],0)+1
provs=[]
for ruc,a in agg.items():
    d,f=duenos.get(ruc,(None,None))
    provs.append({'nombre':a['nombre'],'ruc':ruc,'monto':round(a['monto'],2),
        'n':len(a['convs']),'tipos':a['tipos'],'objeto':a['objeto'],
        'tipo_persona':a['tipo_persona'],'tipo_proveedor_seace':a['tipo_proveedor_seace'],
        'dueno':d,'fuente_dueno':f})
provs.sort(key=lambda x:-x['monto'])
monto_total=round(sum(p['monto'] for p in provs),2)
emp=[p for p in provs if p['tipo_persona']=='empresa']
nat=[p for p in provs if p['tipo_persona']=='natural']
all_convs=set(r[5] for r in rows)
top_personas=sorted([{'nombre':p['nombre'],'ruc':p['ruc'],'monto':p['monto'],'n':p['n']} for p in nat],key=lambda x:-x['monto'])[:20]
out={
 '_meta':{
  'fuente':'OECE/OSCE - CONOSCE Datos Abiertos, reporte de Adjudicaciones (buena pro por item)',
  'fuente_url':'https://conosce.osce.gob.pe/buscador/assets/67ae6c4a/reportes/adjudicaciones/',
  'entidad':'Universidad Nacional de Educacion Enrique Guzman y Valle (UNE - La Cantuta) - pliego 528',
  'ruc':RUC_ENT,
  'codigoentidad_conosce':int(CODIGO),
  'periodo':'2023-2025',
  'extraido':'2026-07',
  'unidad_monto':'Soles (PEN), monto adjudicado por item',
  'nota':"Agregado desde reportes anuales CONOSCE de Adjudicaciones (nivel item de buena pro), archivos CONOSCE_ADJUDICACIONES{2023,2024,2025}_0.xlsx, filtrando codigoentidad 001901 / RUC 20174950971 (UNIVERSIDAD NACIONAL DE EDUCACION ENRIQUE GUZMAN Y VALLE - UNE La Cantuta, pliego 528). 'monto' = suma de monto_adjudicado_item_soles. 'n' = numero de procesos (codigoconvocatoria) distintos en que ese proveedor obtuvo buena pro. 'tipos' = procesos por tipo de proceso de seleccion. 'objeto' = items por objeto contractual (Bien/Servicio/Obra). tipo_persona por prefijo de RUC (20=empresa, 10/15=natural); consorcios (codigo SEACE, no RUC de 11 digitos 10/15/20) clasificados como empresa. NO incluye ordenes de compra <8 UIT (dataset aparte)."
 },
 'totales':{
  'monto_total':monto_total,
  'n_proveedores':len(provs),
  'n_procesos':len(all_convs),
  'n_empresas':len(emp),
  'n_personas_naturales':len(nat),
  'monto_empresas':round(sum(p['monto'] for p in emp),2),
  'monto_personas_naturales':round(sum(p['monto'] for p in nat),2),
 },
 'top_personas':top_personas,
 'proveedores':provs,
}
json.dump(out, open(os.path.join(HERE,'..','data','proveedores-cantuta.json'),'w'),
          ensure_ascii=False, separators=(",",":"))
print("proveedores:", len(provs), "monto_total:", monto_total, "procesos:", len(all_convs))
print("empresas:", len(emp), "naturales:", len(nat))
print("--- TOP 20 ---")
for p in provs[:20]:
    print(f"  {p['monto']:>14,.2f}  {p['ruc']:12}  {p['tipo_persona']:8}  {p['nombre']}")
